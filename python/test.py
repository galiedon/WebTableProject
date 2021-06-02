import warnings

import openpyxl

from const import const
from dbhelper.mysqlHelper import MysqlHelper


def db_test():
    db_helper = MysqlHelper()

    config = {
        'username': const.USERNAME,
        'password': const.PASSWORD,
        'host': const.LOCALHOST,
        'db_name': const.DB_NAME,
    }
    db_helper.connect(**config)

    db_helper.close()


# 0:12:22
def excel_read_test():
    print(const.EXCEL_FILE_PATH)
    with warnings.catch_warnings(record=True):
        warnings.simplefilter('always')

        workbook = openpyxl.load_workbook(const.EXCEL_FILE_PATH)
        sheet = workbook.active

        for row_id in range(sheet.max_row):
            sql = const.INSERT_TABLE_HEAD
            for column in sheet.iter_cols(1, sheet.max_column):
                sql += '"%s", ' % column[row_id].value
                # print(column[0].value)
            dotpos = sql.rfind(',')
            sql = sql[:dotpos]
            sql += const.INSERT_TABLE_END

        workbook.close()


def sql_test():
    sql = 'insert into src_data values("20210508-160106986851", "None", "应明霞[杭州水盾科技有限公司]", "0", "475*400*20-304", ' \
          '"0", "None", "280", "20", "5600", "5600", "按图加工，十套", "20", "0", "20", "9170", "None", "None", "None", ' \
          '"2021年5月8日", "网上银行", "2021年5月8日", "2021年5月23日", "None", "None", "6444", "9170", "9170", "59", "2021年5月8日 ' \
          '16点1分", "2021年5月22日", "杨垒", "杨垒", "None", "0", "None", "0.0000", "9170", "0", "None", "None", "0", "0", ' \
          '"None", "0", "None", "0", "0.0000", "None", "0", "0", "None", "0", "None", "0", "是", "项慧君", "未发货", "59", ' \
          '"0", "59", "None", "None", "否", "0", "None", "None", "None", "None", "None", "0", "销售部", "应明霞,13868151286,' \
          '杭州富阳新登镇金城南路29号", "未续费", "0", "None", "None", "None", "否", "0", "0", "9170", "否"); '

    db_helper = MysqlHelper()
    config = {
        'username': const.USERNAME,
        'password': const.PASSWORD,
        'host': const.LOCALHOST,
        'db_name': const.DB_NAME,
    }
    db_helper.connect(**config)
    # db_helper.exec(const.INSERT_TABLE_SQL, const.INSERT_TABLE_DATA)
    # result = db_helper.query('select * from src_data')
    db_helper.exec(sql)
    # for tmp in result:
    #     print(tmp)
    print(sql)


def read_data_from_excel_to_mysql():
    db_helper = MysqlHelper()
    config = {
        'username': const.USERNAME,
        'password': const.PASSWORD,
        'host': const.LOCALHOST,
        'db_name': const.DB_NAME,
    }
    db_helper.connect(**config)

    with warnings.catch_warnings(record=True):
        warnings.simplefilter('always')

        workbook = openpyxl.load_workbook(const.EXCEL_FILE_PATH)
        sheet = workbook.active

        for row_id in range(sheet.max_row):
            sql = const.INSERT_TABLE_HEAD
            for column in sheet.iter_cols(1, sheet.max_column):
                sql += '"%s", ' % column[row_id].value
                # print(column[0].value)
            dotpos = sql.rfind(',')
            sql = sql[:dotpos]
            sql += const.INSERT_TABLE_END
            # print(sql)
            db_helper.exec(sql)
        workbook.close()
    db_helper.close()


def is_none_data(src, dst):
    if src == 'None':
        return dst

    return src


def read_src_data_2_orders():
    db_helper = MysqlHelper()
    config = {
        'username': const.USERNAME,
        'password': const.PASSWORD,
        'host': const.LOCALHOST,
        'db_name': const.DB_NAME,
    }
    db_helper.connect(**config)

    sql = 'select C1, C3, C5, C8, C9, C10, C20 from src_data;'
    cursor = db_helper.query(sql)
    results = []
    for result in cursor:
        results.append(result)

    tmp = [0, 0, 0, 0, 0, 0, 0]
    none_text = 'None'
    for (order_id,
         customer,
         product_name,
         product_per_price,
         product_num,
         product_sum_price,
         order_time) in results:
        order_id = is_none_data(order_id, tmp[0])
        customer = is_none_data(customer, tmp[1])
        product_name = is_none_data(product_name, tmp[2])
        product_per_price = is_none_data(product_per_price, tmp[3])
        product_num = is_none_data(product_num, tmp[4])
        product_sum_price = is_none_data(product_sum_price, tmp[5])
        order_time = is_none_data(order_time, tmp[6])

        # 处理中文
        # product_num = product_num.replace('个', '').replace('公斤', '')

        # 处理日期格式
        order_time = order_time.replace('年', '-').replace('月', '-').replace('日', '')
        # print(order_time)
        sql = 'insert into orders(order_id, ' \
              'customer, product_name, product_per_price, product_num, product_sum_price, order_time) ' \
              'values("{}","{}","{}", {}, "{}", {}, DATE("{}"));'.format(order_id, customer, product_name,
                                                                         product_per_price, product_num,
                                                                         product_sum_price,
                                                                         order_time)

        db_helper.exec(sql)
        tmp[0] = order_id
        tmp[1] = customer
        tmp[2] = product_name
        tmp[3] = product_per_price
        tmp[4] = product_num
        tmp[5] = product_sum_price
        tmp[6] = order_time

        # print(sql)
        # break
    db_helper.close()
