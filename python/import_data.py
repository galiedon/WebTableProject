import os
import time

import openpyxl
import warnings
from dbhelper.mysqlHelper import MysqlHelper
from const import const
from test import is_none_data

path = "addition.xlsx"
def read_data_from_excel_to_mysql():
    db_helper = MysqlHelper()
    config = {
        'username': const.USERNAME,
        'password': const.PASSWORD,
        'host': const.LOCALHOST,
        'db_name': const.DB_NAME,
    }
    db_helper.connect(**config)
    sql = "drop table if exists data_src_tmp"
    db_helper.exec(sql)
    sql = "create table data_src_tmp(C1 Text, C2 Text, C3 Text, C4 Text, C5 Text, C6 Text, C7 Text, C8 Text, C9 Text, " \
          "C10 Text, C11 Text, C12 Text, C13 Text, C14 Text, C15 Text, C16 Text, C17 Text, C18 Text, C19 Text, " \
          "C20 Text, C21 Text, C22 Text, C23 Text, C24 Text, C25 Text, C26 Text, C27 Text, C28 Text, C29 Text, " \
          "C30 Text, C31 Text, C32 Text, C33 Text, C34 Text, C35 Text, C36 Text, C37 Text, C38 Text, C39 Text, " \
          "C40 Text, C41 Text, C42 Text, C43 Text, C44 Text, C45 Text, C46 Text, C47 Text, C48 Text, C49 Text, " \
          "C50 Text, C51 Text, C52 Text, C53 Text, C54 Text, C55 Text, C56 Text, C57 Text, C58 Text, C59 Text, " \
          "C60 Text, C61 Text, C62 Text, C63 Text, C64 Text, C65 Text, C66 Text, C67 Text, C68 Text, C69 Text, " \
          "C70 Text, C71 Text, C72 Text, C73 Text, C74 Text, C75 Text, C76 Text, C77 Text, C78 Text, C79 Text, " \
          "C80 Text, C81 Text, C82 Text, C83 Text); "
    db_helper.exec(sql)

    file_path = os.path.join(const.RES_PATH, path)

    with warnings.catch_warnings(record=True):
        warnings.simplefilter('always')

        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        for row_id in range(sheet.max_row):
            sql = 'insert into data_src_tmp values('
            for column in sheet.iter_cols(1, sheet.max_column):
                sql += '"%s", ' % column[row_id].value
                # print(column[0].value)
            dotpos = sql.rfind(',')
            sql = sql[:dotpos]
            sql += ')'
            # print(sql)
            db_helper.exec(sql)
        workbook.close()
    db_helper.close()

def read_src_data_2_orders():
    db_helper = MysqlHelper()
    config = {
        'username': const.USERNAME,
        'password': const.PASSWORD,
        'host': const.LOCALHOST,
        'db_name': const.DB_NAME,
    }
    db_helper.connect(**config)

    sql = 'select C1, C3, C5, C8, C9, C10, C20 from data_src_tmp;'
    cursor = db_helper.query(sql)
    results = []
    for result in cursor:
        results.append(result)

    tmp = [0, 0, 0, 0, 0, 0, 0]
    none_text = 'None'
    for (order_id,
         customer,
         product_name,
         product_per_price,
         product_num,
         product_sum_price,
         order_time) in results:
        order_id = is_none_data(order_id, tmp[0])
        customer = is_none_data(customer, tmp[1])
        product_name = is_none_data(product_name, tmp[2])
        product_per_price = is_none_data(product_per_price, tmp[3])
        product_num = is_none_data(product_num, tmp[4])
        product_sum_price = is_none_data(product_sum_price, tmp[5])
        order_time = is_none_data(order_time, tmp[6])

        # 处理中文
        # product_num = product_num.replace('个', '').replace('公斤', '')

        # 处理日期格式
        order_time = order_time.replace('年', '-').replace('月', '-').replace('日', '')
        # print(order_time)
        sql = 'insert into orders(order_id, ' \
              'customer, product_name, product_per_price, product_num, product_sum_price, order_time) ' \
              'values("{}","{}","{}", {}, "{}", {}, DATE("{}"));'.format(order_id, customer, product_name,
                                                                         product_per_price, product_num,
                                                                         product_sum_price,
                                                                         order_time)

        db_helper.exec(sql)
        tmp[0] = order_id
        tmp[1] = customer
        tmp[2] = product_name
        tmp[3] = product_per_price
        tmp[4] = product_num
        tmp[5] = product_sum_price
        tmp[6] = order_time

        # print(sql)
        # break

    db_helper.close()

def importTable(file_path):
    global path
    path = file_path
    read_data_from_excel_to_mysql()
    read_src_data_2_orders()

if __name__ == '__main__':
    start = time.time()
    importTable()
    # excel_read_test()
    # sql_test()
    # read_data_from_excel_to_mysql()
    # read_src_data_2_orders()
    end = time.time()
    delta = end - start
    print('%d:%d:%d' % (delta // 3600, (delta // 60) % 60, delta % 60))
